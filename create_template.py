# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
OUTPUT = BASE / "周维度经营看板上传模板.xlsx"


def sanitize_title(title: str) -> str:
    for char in ["\\", "/", "*", "[", "]", ":", "?"]:
        title = title.replace(char, "-")
    return title[:31]


blue = PatternFill("solid", fgColor="2D6FB8")
light = PatternFill("solid", fgColor="EAF3FC")
white_font = Font(color="FFFFFF", bold=True, size=16)
header_font = Font(color="284D79", bold=True)
thin = Side(style="thin", color="D9E6F2")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_sheet(workbook, title, headers, rows, percent_cols=None, money_cols=None, wrap_cols=None):
    percent_cols = percent_cols or []
    money_cols = money_cols or []
    wrap_cols = wrap_cols or []
    sheet = workbook.create_sheet(sanitize_title(title))

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.fill = light
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.border = all_border
            if col_idx in percent_cols:
                cell.number_format = "0.0%"
            elif col_idx in money_cols:
                cell.number_format = "#,##0.00"
            if col_idx in wrap_cols:
                cell.alignment = Alignment(wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    sheet.freeze_panes = "A2"


workbook = Workbook()
sheet = workbook.active
sheet.title = sanitize_title("填写说明")

sheet.merge_cells("A1:F1")
sheet["A1"] = "周维度经营看板上传模板"
sheet["A1"].fill = blue
sheet["A1"].font = white_font
sheet["A1"].alignment = Alignment(horizontal="center")

instructions = [
    ["说明项", "填写要求"],
    ["周次格式", "建议填写 7.30、2026-W31 或 本周日期"],
    ["首行", "每张表第一行写表头，不要留空"],
    ["金额字段", "填写纯数字，不要手动加逗号"],
    ["比例字段", "填写 0.453 或 45.3% 都可以"],
    ["备注字段", "可为空，用于展示运营说明"],
    ["部署方式", "该模板可配合 GitHub Pages 静态看板使用"],
]

for row_idx, row in enumerate(instructions, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = sheet.cell(row_idx, col_idx, value)
        cell.border = all_border
        if row_idx == 3:
            cell.fill = light
            cell.font = header_font

sheet.column_dimensions["A"].width = 18
sheet.column_dimensions["B"].width = 52

add_sheet(
    workbook,
    "采购数据统计",
    ["周次", "采购合同个数", "采购单个数", "销售采购单个数", "直发采购单个数", "直发采购单比例", "采购金额（元）", "人数", "人均采购金额（元）", "人均采购单（个）", "金额周环比"],
    [
        ["7.09", 1012, 2320, 1100, 488, 0.443, 3980112.2, 15, 265340.81, 154.7, None],
        ["7.16", 1028, 2408, 1126, 502, 0.447, 4042050.4, 15, 269470.03, 160.5, 0.016],
        ["7.23", 1039, 2486, 1169, 521, 0.446, 4141198.82, 15, 276079.92, 165.7, 0.025],
        ["7.30", 1169, 3698, 1138, 515, 0.453, 5760008.07, 15, 384000.54, 246.5, 0.391],
    ],
    percent_cols=[6, 11],
    money_cols=[7, 9, 10],
)

add_sheet(
    workbook,
    "Key-SKU断货统计",
    ["周次", "货号个数", "断货个数", "断货率", "备注说明"],
    [
        ["7.09", 2281, 44, 0.0193, ""],
        ["7.16", 2270, 47, 0.0207, ""],
        ["7.23", 2264, 48, 0.0212, ""],
        ["7.30", 2248, 69, 0.0307, "阈值调整"],
    ],
    percent_cols=[4],
)

add_sheet(
    workbook,
    "议价数据统计",
    ["周次", "单笔议价金额（元）", "备货议价金额（元）", "议价合计（元）"],
    [
        ["7.09", 4880.2, 50210.4, 55090.6],
        ["7.16", 4955.4, 54105.2, 59060.6],
        ["7.23", 4971.97, 56502.88, 61474.85],
        ["7.30", 5210.54, 25998.03, 31208.57],
    ],
    money_cols=[2, 3, 4],
)

add_sheet(
    workbook,
    "仓储单位产出",
    ["月份", "主仓单位面积销售额", "廊坊仓单位面积销售额", "广州仓单位面积销售额", "成都仓单位面积销售额", "主仓单位面积毛利", "廊坊仓单位面积毛利", "广州仓单位面积毛利", "成都仓单位面积毛利"],
    [
        ["2026-06", 1814, 1092, 519, 558, 544, 290, 184, 198],
        ["2026-07", 1850, 1110, 536, 572, 552, 301, 192, 205],
    ],
    money_cols=[2, 3, 4, 5, 6, 7, 8, 9],
)

add_sheet(
    workbook,
    "效率提升",
    ["周次", "模块", "本周进展"],
    [
        ["7.23", "AI", "知识库：产品资料投喂；发票命名自动化"],
        ["7.30", "流程优化", "年度协议管理打通；标签自动流转模块沟通"],
    ],
    wrap_cols=[3],
)

add_sheet(
    workbook,
    "库存周转分析",
    ["仓库/品类", "6月库存金额", "6月平均周转天数", "7月库存金额", "7月平均周转天数", "库存金额环比变化", "周转天数变化"],
    [
        ["主仓有阈值的明星产品", 12960913.76, 79.1, 14129745.66, 90.1, 0.09, 11],
        ["主仓所有商品", 32900585.42, 89.1, 34198877.49, 94.7, 0.039, 5.6],
        ["成都所有商品", 1979750.45, 50.5, 1851979.97, 50.6, -0.065, 0.1],
    ],
    percent_cols=[6],
    money_cols=[2, 3, 4, 5, 7],
)

workbook.save(OUTPUT)
print(OUTPUT)
